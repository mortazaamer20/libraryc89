from __future__ import annotations

import io
from datetime import timedelta

from django.conf import settings
from django.db.models import Count, Q
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from bidi.algorithm import get_display
import arabic_reshaper

from api.models import Book, BookCopy, Loan, LibrarySection, Author


def _reshape_arabic(text: str) -> str:
    if not text:
        return ""
    return get_display(arabic_reshaper.reshape(str(text)))


def resolve_period(period: str, date_from=None, date_to=None):
    today = timezone.localdate()

    if period == "today":
        return today, today

    if period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end

    if period == "month":
        start = today.replace(day=1)
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        return start, end

    if period == "custom":
        return date_from, date_to

    return None, None


def build_report_data(period: str, date_from=None, date_to=None):
    start_date, end_date = resolve_period(period, date_from, date_to)

    books_qs = Book.objects.select_related("section", "shelf").prefetch_related("authors", "copies", "files")
    loans_qs = Loan.objects.select_related("copy", "copy__book", "borrower", "handled_by")
    sections_qs = LibrarySection.objects.select_related("parent")
    authors_qs = Author.objects.all()
    copies_qs = BookCopy.objects.select_related("book", "shelf", "book__section")

    if start_date and end_date:
        books_added = books_qs.filter(created_at__date__range=(start_date, end_date))
        loans_created = loans_qs.filter(loan_date__range=(start_date, end_date))
        returned_loans = loans_qs.filter(
            status=Loan.Status.RETURNED,
            returned_at__range=(start_date, end_date),
        )
        sections_added = sections_qs.filter(created_at__date__range=(start_date, end_date))
        authors_added = authors_qs.filter(created_at__date__range=(start_date, end_date))
    else:
        books_added = books_qs.all()
        loans_created = loans_qs.all()
        returned_loans = loans_qs.filter(status=Loan.Status.RETURNED)
        sections_added = sections_qs.all()
        authors_added = authors_qs.all()

    active_loans = loans_qs.filter(status=Loan.Status.ACTIVE)
    overdue_loans = loans_qs.filter(status=Loan.Status.ACTIVE, due_date__lt=timezone.localdate())

    copies_status_counts = (
        copies_qs.values("status")
        .annotate(total=Count("id"))
        .order_by("status")
    )

    sections_counts = (
        sections_qs.annotate(books_count=Count("books", distinct=True))
        .order_by("name")
    )

    authors_counts = (
        authors_qs.annotate(books_count=Count("books", distinct=True))
        .order_by("full_name")
    )

    books_detail = books_added.annotate(
        copies_count=Count("copies", distinct=True),
        files_count=Count("files", distinct=True),
    ).order_by("-created_at")

    loans_detail = loans_created.order_by("-loan_date")
    returns_detail = returned_loans.order_by("-returned_at")
    overdue_detail = overdue_loans.order_by("due_date")

    summary = {
        "sections_total": sections_qs.count(),
        "sections_added": sections_added.count(),
        "authors_total": authors_qs.count(),
        "authors_added": authors_added.count(),
        "books_total": books_qs.count(),
        "books_added": books_added.count(),
        "copies_total": copies_qs.count(),
        "active_loans": active_loans.count(),
        "loans_created": loans_created.count(),
        "returns_count": returned_loans.count(),
        "overdue_count": overdue_loans.count(),
    }

    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "summary": summary,
        "copies_status_counts": list(copies_status_counts),
        "sections_counts": list(sections_counts),
        "authors_counts": list(authors_counts),
        "books_detail": list(books_detail),
        "loans_detail": list(loans_detail),
        "returns_detail": list(returns_detail),
        "overdue_detail": list(overdue_detail),
    }


def export_excel_report(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "الملخص"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_table_style(sheet, start_row, start_col, end_row, end_col):
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Summary sheet
    ws.merge_cells("A1:B1")
    ws["A1"] = "تقرير المكتبة"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["البند", "العدد"])
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    summary = data["summary"]
    summary_rows = [
        ("إجمالي الأقسام", summary["sections_total"]),
        ("الأقسام المضافة داخل الفترة", summary["sections_added"]),
        ("إجمالي المؤلفين", summary["authors_total"]),
        ("المؤلفون المضافون داخل الفترة", summary["authors_added"]),
        ("إجمالي الكتب", summary["books_total"]),
        ("الكتب المضافة داخل الفترة", summary["books_added"]),
        ("إجمالي النسخ", summary["copies_total"]),
        ("الإعارات السارية", summary["active_loans"]),
        ("الإعارات المنشأة داخل الفترة", summary["loans_created"]),
        ("الكتب المسترجعة داخل الفترة", summary["returns_count"]),
        ("الكتب المتأخرة", summary["overdue_count"]),
    ]
    for label, value in summary_rows:
        ws.append([label, value])

    apply_table_style(ws, 2, 1, ws.max_row, 2)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A3"

    def add_sheet(title, headers, rows):
        sheet = wb.create_sheet(title=title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append(row)

        apply_table_style(sheet, 1, 1, sheet.max_row, len(headers))
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 24
        sheet.freeze_panes = "A2"
        return sheet

    add_sheet(
        "الكتب",
        ["العنوان", "القسم", "الرف", "ISBN", "رقم الأرشفة", "التاريخ"],
        [
            [
                book.title,
                str(book.section),
                str(book.shelf) if book.shelf else "",
                book.isbn,
                book.archive_number,
                book.created_at.strftime("%Y-%m-%d"),
            ]
            for book in data["books_detail"]
        ],
    )

    add_sheet(
        "الإعارات",
        ["الكتاب", "النسخة", "المستعير", "الحالة", "تاريخ الإعارة", "الموعد"],
        [
            [
                loan.copy.book.title,
                loan.copy.copy_number,
                str(loan.borrower),
                loan.status,
                loan.loan_date.strftime("%Y-%m-%d") if loan.loan_date else "",
                loan.due_date.strftime("%Y-%m-%d") if loan.due_date else "",
            ]
            for loan in data["loans_detail"]
        ],
    )

    add_sheet(
        "المسترجع",
        ["الكتاب", "النسخة", "المستعير", "تاريخ الإعارة", "تاريخ الإرجاع"],
        [
            [
                loan.copy.book.title,
                loan.copy.copy_number,
                str(loan.borrower),
                loan.loan_date.strftime("%Y-%m-%d") if loan.loan_date else "",
                loan.returned_at.strftime("%Y-%m-%d") if loan.returned_at else "",
            ]
            for loan in data["returns_detail"]
        ],
    )

    add_sheet(
        "المتأخر",
        ["الكتاب", "النسخة", "المستعير", "تاريخ الاستحقاق"],
        [
            [
                loan.copy.book.title,
                loan.copy.copy_number,
                str(loan.borrower),
                loan.due_date.strftime("%Y-%m-%d") if loan.due_date else "",
            ]
            for loan in data["overdue_detail"]
        ],
    )

    add_sheet(
        "الأقسام",
        ["القسم", "القسم الأب", "عدد الكتب"],
        [
            [
                section.name,
                section.parent.name if section.parent else "",
                section.books_count,
            ]
            for section in data["sections_counts"]
        ],
    )

    add_sheet(
        "المؤلفون",
        ["المؤلف", "الكتب", "البلد"],
        [
            [
                author.full_name,
                author.books_count,
                author.country,
            ]
            for author in data["authors_counts"]
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf_report(data: dict, font_path: str | None = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ArabicTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica",
        fontSize=18,
        leading=24,
    )

    normal_style = ParagraphStyle(
        "ArabicNormal",
        parent=styles["BodyText"],
        alignment=TA_RIGHT,
        fontName="Helvetica",
        fontSize=10,
        leading=14,
    )

    if font_path:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        pdfmetrics.registerFont(TTFont("ArabicFont", font_path))
        title_style.fontName = "ArabicFont"
        normal_style.fontName = "ArabicFont"

    story.append(Paragraph(_reshape_arabic("تقرير المكتبة"), title_style))
    story.append(Spacer(1, 8))

    period_text = "الفترة: "
    if data["start_date"] and data["end_date"]:
        period_text += f"{data['start_date']} → {data['end_date']}"
    else:
        period_text += "الكلي"

    story.append(Paragraph(_reshape_arabic(period_text), normal_style))
    story.append(Spacer(1, 10))

    summary = data["summary"]
    summary_table_data = [
        [_reshape_arabic("البند"), _reshape_arabic("العدد")],
        [_reshape_arabic("إجمالي الأقسام"), summary["sections_total"]],
        [_reshape_arabic("الأقسام المضافة داخل الفترة"), summary["sections_added"]],
        [_reshape_arabic("إجمالي المؤلفين"), summary["authors_total"]],
        [_reshape_arabic("المؤلفون المضافون داخل الفترة"), summary["authors_added"]],
        [_reshape_arabic("إجمالي الكتب"), summary["books_total"]],
        [_reshape_arabic("الكتب المضافة داخل الفترة"), summary["books_added"]],
        [_reshape_arabic("إجمالي النسخ"), summary["copies_total"]],
        [_reshape_arabic("الإعارات السارية"), summary["active_loans"]],
        [_reshape_arabic("الإعارات المنشأة داخل الفترة"), summary["loans_created"]],
        [_reshape_arabic("الكتب المسترجعة داخل الفترة"), summary["returns_count"]],
        [_reshape_arabic("الكتب المتأخرة"), summary["overdue_count"]],
    ]

    tbl = Table(summary_table_data, colWidths=[120 * mm, 30 * mm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 12))

    def add_section(title: str, headers: list[str], rows: list[list[str]]):
        story.append(Paragraph(_reshape_arabic(title), title_style))
        story.append(Spacer(1, 6))
        table_data = [[_reshape_arabic(h) for h in headers]]
        table_data.extend([[str(cell) for cell in row] for row in rows])
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 10))

    add_section(
        "الكتب المضافة",
        ["العنوان", "القسم", "الرف", "ISBN", "الأرشفة", "التاريخ"],
        [
            [
                book.title,
                str(book.section),
                str(book.shelf) if book.shelf else "",
                book.isbn,
                book.archive_number,
                book.created_at.strftime("%Y-%m-%d"),
            ]
            for book in data["books_detail"][:25]
        ],
    )

    add_section(
        "الإعارات",
        ["الكتاب", "النسخة", "المستعير", "الحالة", "الإعارة", "الموعد"],
        [
            [
                loan.copy.book.title,
                loan.copy.copy_number,
                str(loan.borrower),
                loan.status,
                loan.loan_date.strftime("%Y-%m-%d") if loan.loan_date else "",
                loan.due_date.strftime("%Y-%m-%d") if loan.due_date else "",
            ]
            for loan in data["loans_detail"][:25]
        ],
    )

    add_section(
        "المسترجع",
        ["الكتاب", "النسخة", "المستعير", "الإعارة", "الإرجاع"],
        [
            [
                loan.copy.book.title,
                loan.copy.copy_number,
                str(loan.borrower),
                loan.loan_date.strftime("%Y-%m-%d") if loan.loan_date else "",
                loan.returned_at.strftime("%Y-%m-%d") if loan.returned_at else "",
            ]
            for loan in data["returns_detail"][:25]
        ],
    )

    def add_page_number(canvas, doc):
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(200 * mm, 10 * mm, f"Page {doc.page}")

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return buffer.getvalue()